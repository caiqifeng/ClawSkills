from concurrent.futures import ThreadPoolExecutor, as_completed
import json
from datetime import datetime, timedelta
import os
from typing import Dict, List, Set, Tuple

import openpyxl
import requests


def custom_sort_key(item):
    key_priority = {"PC": 0, "Xbox": 1, "PS5": 2}  # 优先级规则
    return key_priority.get(item[0], 999)


def get_taskIdList(pipelineIdList: list, startTimeAfter: str, endTime: str):
    # 通过流水线ID与指定时间动态获取任务ID数组
    taskIdList = []
    for pipelineId in pipelineIdList:
        res_json = {
            "projectId": projectId,
            "order_by": "queueTime",
            "asc": False,
            "filters": {
                "pipelineId": pipelineId,
                "startTime": startTimeAfter,
                "endTime": endTime,
            },
            "page": 1,
            "count": 20,
        }
        tasklist = requests.post(
            f"https://automation-api.testplus.cn/api/tasks/list?projectId={projectId}",
            json=res_json,
        ).json()["data"]["list"]
        for task in tasklist:
            buildId = task["buildId"]
            taskIdList.append(buildId)
    # print(taskIdList)
    return taskIdList
def get_taskInfo_by_taskId(taskIdList: list, startTimeAfter):
    # 通过任务ID与指定时间动态获取任务ID数组
    taskIdDict = {}
    for taskId in taskIdList:
        taskdata = requests.get(
            f"https://automation-api.testplus.cn/api/tasks/detail/{taskId}?projectId={projectId}"
        ).json()["data"]
        createTime = taskdata["createTime"].replace("T", " ")
        dt_createTime = datetime.strptime(createTime, "%Y-%m-%d %H:%M:%S")
        dt_startTimeAfter = datetime.strptime(startTimeAfter, "%Y-%m-%d %H:%M:%S")
        if dt_createTime < dt_startTimeAfter:
            print(
                f"任务ID: {taskId} 创建时间: {createTime} 在指定时间{startTimeAfter}之前，跳过"
            )
            continue

        packageVersion = taskdata["packageVersion"]
        model_data = json.loads(taskdata["model"])
        platform = model_data["baseInfo"]["platform"]
        if platform not in taskIdDict:
            taskIdDict[platform] = {}

        if packageVersion not in taskIdDict[platform]:
            taskIdDict[platform][packageVersion] = []

        taskinfo = {}
        keys_to_keep = ["buildId", "buildName", "caseDetails"]
        new_taskdata = {key: taskdata[key] for key in keys_to_keep if key in taskdata}
        taskIdDict[platform][packageVersion].append(new_taskdata)

    # print(taskIdDict)
    return taskIdDict


def find_case_ex_by_id(perfeyeID: str) -> List[str]:
    """单个 perfeyeID 的处理逻辑"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.66 Safari/537.36 Edg/103.0.1264.44",
        "Authorization": "Bearer mj6cltF&!L#yWX8k",
    }
    test_url = f"https://perfeye.testplus.cn/api/show/task/{perfeyeID}"
    response_data = requests.post(url=test_url, headers=headers, stream=False)
    response_data = response_data.json()

    seen_ids: Set[str] = set()
    unique_list: List[str] = []

    for item in response_data.get("data", {}).get("DataList", []):
        item_id = item.get("label")
        if item_id and item_id not in seen_ids:
            seen_ids.add(item_id)
            unique_list.append(item_id)

    record_case = []
    for index, item in enumerate(unique_list):
        if "流程异常" in item:
            record_case.append(unique_list[index - 1] + "流程异常")
        if "失败" in item:
            record_case.append(unique_list[index] + "关卡失败")

    return record_case


def process_device_task(
    version: str, device: str, perfeye_id: str
) -> Tuple[str, str, str, List[str]]:
    """单个设备的处理任务"""
    result = find_case_ex_by_id(perfeye_id)
    return version, device, perfeye_id, result


def export_to_excel_with_sheets(
    all_results: List[Dict], filename: str, sheet_name: str = "稳定性汇总"
):
    """导出到单一 Sheet，所有平台数据合并（带平台字段）"""

    # 确保文件名有 .xlsx 后缀
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    # 检查文件是否存在
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()
        # 删除默认的 sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

    # 如果 sheet 已存在，直接使用（不删除）
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 追加数据到现有数据后面
        start_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(title=sheet_name)
        start_row = 1

        # 表头（只在第一次创建时写入）
        headers = [
            "平台",
            "版本",
            "设备",
            "PerfeyeID",
            "Perfeye 链接",
            "状态",
            "异常数量",
            "异常详情",
        ]
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="4472C4", fill_type="solid"
            )
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        # 设置列宽
        ws.column_dimensions["A"].width = 12  # 平台
        ws.column_dimensions["B"].width = 15  # 版本
        ws.column_dimensions["C"].width = 40  # 设备
        ws.column_dimensions["D"].width = 40  # PerfeyeID
        ws.column_dimensions["E"].width = 50  # Perfeye 链接
        ws.column_dimensions["F"].width = 12  # 状态
        ws.column_dimensions["G"].width = 12  # 异常数量
        ws.column_dimensions["H"].width = 80  # 异常详情

        start_row = 2

    # 写入数据
    for row in all_results:
        ws.append(
            [
                row["平台"],
                row["版本"],
                row["设备"],
                row["PerfeyeID"],
                row["Perfeye 链接"],
                row["状态"],
                row["异常数量"],
                row["异常详情"],
            ]
        )

    # 状态列条件格式（只对新添加的行）
    for row in range(start_row, start_row + len(all_results)):
        cell = ws[f"F{row}"]
        if cell.value == "存在异常":
            cell.fill = openpyxl.styles.PatternFill(
                start_color="FFC7CE", fill_type="solid"
            )
            cell.font = openpyxl.styles.Font(color="9C0006")
        else:
            cell.fill = openpyxl.styles.PatternFill(
                start_color="C6EFCE", fill_type="solid"
            )
            cell.font = openpyxl.styles.Font(color="006100")

    # 将链接列设置为超链接格式
    for row in range(start_row, start_row + len(all_results)):
        cell = ws[f"E{row}"]
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    wb.save(filename)
    print(
        f"📁 结果已导出到：{filename} (Sheet: {sheet_name}, 当前总行数：{ws.max_row})"
    )


def create_summary_sheet(filename: str):
    """创建统计摘要 Sheet（按平台分组统计）"""
    wb = openpyxl.load_workbook(filename)

    # 如果已有统计摘要 Sheet，删除重建
    if "统计摘要" in wb.sheetnames:
        del wb["统计摘要"]

    ws_summary = wb.create_sheet(title="统计摘要", index=0)  # 放在第一个位置

    # 从"稳定性汇总"Sheet 读取数据
    if "稳定性汇总" not in wb.sheetnames:
        print("⚠️  未找到'稳定性汇总'Sheet，跳过统计摘要生成")
        return

    ws = wb["稳定性汇总"]

    # 按平台分组统计（平台在第 1 列，状态在第 6 列）
    summary_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # 跳过空行
            continue
        platform = row[0]  # 平台列
        status = row[5]  # 状态列

        if platform not in summary_data:
            summary_data[platform] = {"total": 0, "abnormal": 0, "normal": 0}

        summary_data[platform]["total"] += 1
        if status == "存在异常":
            summary_data[platform]["abnormal"] += 1
        else:
            summary_data[platform]["normal"] += 1

    # 写入摘要
    ws_summary.append(["平台", "总设备数", "存在异常", "无异常", "异常率"])

    for platform, data in summary_data.items():
        abnormal_rate = (
            f"{data['abnormal'] / data['total'] * 100:.2f}%"
            if data["total"] > 0
            else "0%"
        )
        ws_summary.append(
            [platform, data["total"], data["abnormal"], data["normal"], abnormal_rate]
        )

    # 总计
    total_all = sum(d["total"] for d in summary_data.values())
    abnormal_all = sum(d["abnormal"] for d in summary_data.values())
    normal_all = total_all - abnormal_all
    abnormal_rate_all = (
        f"{abnormal_all / total_all * 100:.2f}%" if total_all > 0 else "0%"
    )

    ws_summary.append([])
    ws_summary.append(["总计", total_all, abnormal_all, normal_all, abnormal_rate_all])

    # 设置样式
    for cell in ws_summary[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 12
    ws_summary.column_dimensions["D"].width = 12
    ws_summary.column_dimensions["E"].width = 12

    wb.save(filename)
    print(f"📊 统计摘要已生成")


def check_all_devices_multithread(
    version_device_perfeye_dic: Dict[str, Dict[str, List[str]]],
    platform: str,
    max_workers: int = 10,
    output_excel: str = None,
    sheet_name: str = "稳定性汇总",
) -> Dict[str, Dict[str, List[str]]]:
    """
    多线程批量检查所有设备的 perfeye 数据，并写入 Excel

    Args:
        version_device_perfeye_dic: {version: {device: [perfeye_id1, perfeye_id2, ...]}}
        platform: 平台名称（PC、Android、iOS 等）
        max_workers: 最大线程数
        output_excel: 输出 Excel 文件名（不含.xlsx）
        sheet_name: Sheet 名称

    Returns:
        更新后的字典（只保留有异常的设备）
    """
    tasks = []
    print(version_device_perfeye_dic)
    
    for version, device_perfeye_dic in version_device_perfeye_dic.items():
        for device, perfeye_ids in device_perfeye_dic.items():
            # 兼容处理：如果是字符串则转为列表
            if isinstance(perfeye_ids, str):
                perfeye_ids = [perfeye_ids]
            
            # 遍历每个 perfeye_id
            for perfeye_id in perfeye_ids:
                tasks.append((version, device, perfeye_id))

    print(f"📋 [{platform}] 共 {len(tasks)} 个 Perfeye 任务需要检查...\n")

    results = {}
    all_results = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_task = {
            executor.submit(process_device_task, v, d, u): (v, d, u)
            for v, d, u in tasks
        }

        for future in as_completed(future_to_task):
            version, device, perfeye_id, result = future.result()

            perfeye_url = f"https://perfeye.testplus.cn/case/{perfeye_id}/report"

            if result:
                print(f"⚠️  [{version}] {device} 存在异常 ({len(result)} 个) - PerfeyeID: {perfeye_id}")

                if version not in results:
                    results[version] = {}
                if device not in results[version]:
                    results[version][device] = []
                # 合并同一设备的多个 perfeye_id 的异常
                results[version][device].extend(result)

                all_results.append(
                    {
                        "平台": platform,
                        "版本": version,
                        "设备": device,
                        "PerfeyeID": perfeye_id,
                        "Perfeye 链接": perfeye_url,
                        "状态": "存在异常",
                        "异常数量": len(result),
                        "异常详情": "；".join(result),
                    }
                )
            else:
                print(f"✅ [{version}] {device} 无异常 - PerfeyeID: {perfeye_id}")

                all_results.append(
                    {
                        "平台": platform,
                        "版本": version,
                        "设备": device,
                        "PerfeyeID": perfeye_id,
                        "Perfeye 链接": perfeye_url,
                        "状态": "无异常",
                        "异常数量": 0,
                        "异常详情": "",
                    }
                )

    print(f"📊 [{platform}] 检查完成：{len(results)} 个设备存在异常\n")

    if output_excel:
        export_to_excel_with_sheets(all_results, output_excel, sheet_name)

    return results


def get_checkpoint_error(taskID, buildCaseId, deviceId, projectId):
    res = requests.get(
        f"https://automation-api.testplus.cn/api/tasks/device/execute/info?taskId={taskID}&buildCaseId={buildCaseId}&deviceId={deviceId}&projectId={projectId}"
    ).json()["data"][0]["executeData"]
    for item in res[::-1]:
        if (
            "执行失败" in item["msg"]
            and "事件" in item["msg"]
            and "稳定性" not in item["msg"]
        ):
            msg = item["msg"].split("@@")[0].replace("事件：", "")
            if "执行失败" in msg:
                msg = msg.replace("执行失败", "出现")
            return msg
        elif (
            "stack" in item
            and item["stack"]
            and "游戏启动失败或设备掉线" in item["stack"]
        ):
            return "游戏启动流程出现"
        elif (
            "stack" in item
            and item["stack"]
            and "游戏初始化流程出现宕机" in item["stack"]
        ):
            return "游戏初始化流程出现"
    return ""


def get_task_msg(taskIdDict: dict, Regulation_hours: int,output_excel: str = None):
    msg = ""
    index = 0
    INDEX_LIST = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
    for platform, versionDict in taskIdDict.items():
        msg += f"\n\n**{INDEX_LIST[index]}、{platform}**"
        index += 1
        version_msg = ""
        versionMachineCount = 0
        version_run_enough_device = 0
        version_device_perfeye_dic = {}
        device_perfeye_dic = {}
        for version, taskInfoList in versionDict.items():
            version_msg += f"\n\n - （版本：v{version}）\n"
            taskIndex = 1
            for taskInfo in taskInfoList:
                taskID = taskInfo["buildId"]
                taskURL = f"https://uauto2.testplus.cn/project/{projectId}/taskDetail?taskId={taskID}"
                taskresult = f"[{taskInfo['buildName']}]({taskURL})"
                # device_perfeye_dic  = {version:[]}
                
                version_msg += f"\n\t - {taskIndex}.{taskresult} 任务执行汇总："
                taskIndex += 1

                caseDetails = taskInfo["caseDetails"]
                for caseInfo in caseDetails:
                    Abnormal_Info = {}
                    if StabilityName in caseInfo["caseName"]:
                        buildCaseId = caseInfo["buildCaseId"]
                        machineCount = 0
                        deviceDetail = caseInfo["deviceDetail"]
                        run_enough_device = 0
                        for deviceInfo in deviceDetail:
                            if "startTime" in deviceInfo and deviceInfo["startTime"]:
                                startTime = datetime.strptime(
                                    deviceInfo["startTime"], "%Y-%m-%dT%H:%M:%S"
                                )
                                if deviceInfo["endTime"]:
                                    endTime = datetime.strptime(
                                        deviceInfo["endTime"], "%Y-%m-%dT%H:%M:%S"
                                    )
                                else:
                                    # 设备任务仍在运行中的，以当前时间作为endTime
                                    endTime = datetime.now()

                                total_seconds = (endTime - startTime).total_seconds()

                                hours = int(total_seconds // 3600)
                                minutes = int((total_seconds % 3600) // 60)
                                seconds = int(total_seconds % 60)

                                formatted_time = f"{hours}小时{minutes}分{seconds}秒"

                                if total_seconds >= Regulation_hours * 3600:
                                    run_enough_device += 1
                                if total_seconds > 0:
                                    machineCount += 1
                                device_error_type = []
                                perfeye_url = ""

                                if (
                                    "reportData" in deviceInfo
                                    and deviceInfo["reportData"]
                                ):
                                    reportData = deviceInfo["reportData"]
                                    device = f"{deviceInfo['deviceName']}（{deviceInfo['ip']}）"
                                    for report, url in reportData.items():
                                        print(
                                            f"[DEBUG] platform={platform}, device={device}, report key={report}"
                                        )
                                        if "perfeye" in report.lower():
                                            perfeye_url = f"https://perfeye.testplus.cn/case/{url}/report?appKey={projectId}"
                                            if device not in device_perfeye_dic:
                                                device_perfeye_dic[device] = [url]
                                            else:
                                                device_perfeye_dic[device].append(url)
                                            # device_perfeye_dic.update({device: url})
                                        if "Crasheye" in report:
                                            if device not in Abnormal_Info:
                                                Abnormal_Info[device] = {}
                                            Abnormal_Info[device]["Crasheye"] = url
                                            device_error_type.append("宕机")
                                        # 安卓数据类型
                                        if "闪退" in report:
                                            if device not in Abnormal_Info:
                                                Abnormal_Info[device] = {}
                                            Abnormal_Info[device]["Crasheye"] = url
                                            device_error_type.append("闪退")
                                        # 旧版数据类型
                                        if "crasheyeDumpKeys" in report:
                                            print(reportData["crasheyeDumpKeys"])

                                            if device not in Abnormal_Info:
                                                Abnormal_Info[device] = {}

                                                def get_date_str(days: int):
                                                    # 获取日期为指定格式(例：2025-10-17)
                                                    current_date = datetime.now()
                                                    one_month_later = (
                                                        current_date
                                                        + timedelta(days=days)
                                                    )
                                                    formatted_date = (
                                                        one_month_later.strftime(
                                                            "%Y-%m-%d"
                                                        )
                                                    )
                                                    return formatted_date

                                            url = rf"https://crasheye2.testplus.cn/project/{projectId}/vk/{reportData['appkey']}/error?startTime/={get_date_str(0)}&endTime={get_date_str(7)}&searchs=dump_key%7Ci%7C{reportData['crasheyeDumpKeys'][0]}"

                                            print(url)
                                            Abnormal_Info[device]["Crasheye"] = url
                                            device_error_type.append("宕机")
                                        if "卡死" in report:
                                            if device not in Abnormal_Info:
                                                Abnormal_Info[device] = {}
                                            if url:
                                                Abnormal_Info[device]["Dump文件"] = url
                                            device_error_type.append("卡死")
                                        if "GpuDump" in report:
                                            if device not in Abnormal_Info:
                                                Abnormal_Info[device] = {}
                                            Abnormal_Info[device]["GPUDump"] = url
                                            device_error_type.append("GPU宕机")
                                    if device_error_type:
                                        for log in deviceInfo["logUrl"]:
                                            if (
                                                "Player.log" in log
                                                or "Player.zip" in log
                                            ):
                                                Abnormal_Info[device]["游戏日志"] = log
                                                break
                                            if "game-log" in log:
                                                Abnormal_Info[device]["游戏日志"] = log
                                                break
                                            if "system-log" in log:
                                                Abnormal_Info[device][
                                                    "系统日志"
                                                ] = log

                                        if perfeye_url:
                                            Abnormal_Info[device]["Perfeye"] = (
                                                perfeye_url
                                            )

                                        deviceId = deviceInfo["deviceId"]
                                        checkpoint_error = get_checkpoint_error(
                                            taskID, buildCaseId, deviceId, projectId
                                        )
                                        for error_type in device_error_type:
                                            print(checkpoint_error)
                                            print(error_type)
                                            checkpoint_error += error_type
                                            if error_type != device_error_type[-1]:
                                                checkpoint_error += "+"
                                        Abnormal_Info[device]["checkpoint_error"] = (
                                            checkpoint_error
                                        )
                                        Abnormal_Info[device]["时长"] = formatted_time
                        version_msg += f"共**{machineCount}台**设备，其中**{run_enough_device}台**设备执行超过**{Regulation_hours}小时**，"
                        versionMachineCount += machineCount
                        version_run_enough_device += run_enough_device
                        Abnormal_Num = len(Abnormal_Info)
                        if Abnormal_Num > 0:
                            version_msg += f"<font color='red'>以下{Abnormal_Num}台出现卡死和宕机</font>"
                        else:
                            version_msg += "<font color='green'>未发现卡死和宕机</font>"

                        for device, abnormalList in Abnormal_Info.items():
                            print(Abnormal_Info)
                            # with open("./data.txt", "w", encoding="utf-8") as f:
                            #     f.write(str(Abnormal_Info))
                            # if find_case_ex_by_id
                            RunTime = Abnormal_Info[device]["时长"]
                            checkpoint_error = Abnormal_Info[device]["checkpoint_error"]
                            version_msg += f"\n\t\t  - {device}执行了{RunTime}，在{checkpoint_error} "
                            Abnormal_Info[device].pop("时长")
                            Abnormal_Info[device].pop("checkpoint_error")
                            for abnormal, url in Abnormal_Info[device].items():
                                version_msg += f"[{abnormal}]({url})"
                                if (
                                    len(Abnormal_Info[device].keys()) > 1
                                    and abnormal
                                    != list(Abnormal_Info[device].keys())[-1]
                                ):
                                    version_msg += " | "
                print(device_perfeye_dic)
            print("xixi")
            print(device_perfeye_dic)
            version_device_perfeye_dic.update({version: device_perfeye_dic})
        print("haha")
        print(version_device_perfeye_dic)
        msg += f"（共**{versionMachineCount}台**设备，其中**{version_run_enough_device}台**执行超过**{Regulation_hours}小时**）"
        msg += version_msg
        print("记录案例流程异常信息")
        if output_excel:
            version_device_perfeye_dic = check_all_devices_multithread(
                version_device_perfeye_dic,
                platform=platform,
                max_workers=10,
                output_excel=output_excel,
                sheet_name="稳定性汇总",
            )
            print("xixi2")
            print(version_device_perfeye_dic)

    print(msg)
    return msg


if __name__ == "__main__":
    projectId = "jxsj4"
    Regulation_hours = 4  # 统计超出统计时长的数据
    StabilityName = "稳定性"  # 监控的案例名
    taskIdList = []
    output_excel = "F:\\Downloads\\stability_summary" # 输出 Excel 文件名（不含.xlsx）

    # 删除旧的 Excel 文件，避免数据累积
    if os.path.exists(output_excel + ".xlsx"):
        os.remove(output_excel + ".xlsx")
        print(f"🗑️  已删除旧的 {output_excel}.xlsx 文件")
    # 1 获取指定时间后的任务数据
    startTimeAfter = "2026-04-10"  # 汇总这个时间段后的数据，例：2026-01-12 15:06:00 2026-01-30 21:00:00
    if not startTimeAfter:
        # 如果没有指定开始时间，则默认为当天0点
        start_of_day = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        startTimeAfter = start_of_day.strftime("%Y-%m-%d %H:%M:%S")

    if ":" not in startTimeAfter:  # 如果没有指定时间，则默认为当天0点
        startTimeAfter = startTimeAfter + " 00:00:00"

    endTime = "2026-04-10 23:59:59"
    if not endTime:
        # 如果没有指定截止时间，则默认为当前时间
        endTime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 2 可获取指定流水线ID下的任务
    pipelineIdList = [263, 983, 466, 917, 649]  # 指定流水线
    if pipelineIdList:
        # 不指定任务ID时，可通过流水线ID与指定时间动态获取任务ID列表
        taskIdList = get_taskIdList(pipelineIdList, startTimeAfter, endTime)
    
    # 3 不指定流水线时，可直接指定具体的任务ID，适用于临时任务
    extend_taskIdList = []  # 直接指定任务ID列表

    taskIdList.extend(extend_taskIdList)

    # 4. 删除taskIdList中的任务ID
    delete_taskIdList = []
    taskIdList = [task for task in taskIdList if task not in delete_taskIdList]

    taskInfo = get_taskInfo_by_taskId(taskIdList, startTimeAfter)
    taskInfo = dict(sorted(taskInfo.items(), key=custom_sort_key))
    print(taskInfo)
    # ,'output_excel'
    msg = get_task_msg(taskInfo, Regulation_hours)
    Title = f"#### **{startTimeAfter.split(' ')[0].replace('-', '.')}《剑侠世界4》稳定性汇总**\n"
    All_Summary = Title + msg
    filename = rf"Stability_Summary.md"
    with open(filename, "w", encoding="utf-8") as file:
        file.write(All_Summary)
